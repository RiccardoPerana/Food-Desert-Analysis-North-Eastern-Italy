# Exports the final results list to a clean, formatted .xlsx spreadsheet.
# Rows flagged for manual review (distance >= config.DISTANCE_REVIEW_THRESHOLD_KM) are highlighted in yellow.

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def export_to_spreadsheet(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Underserved Towns"

    headers = ["Town", "Province", "Population", "Distance to Nearest Supermarket (km)",
               "Nearest Supermarket", "Flagged for Review"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    flag_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    sorted_results = sorted(results, key=lambda r: r["distance_km"], reverse=True)

    for r in sorted_results:
        flagged = r.get("flagged_for_review", False)
        ws.append([
            r["name"],
            r["province"],
            r["population"],
            r["distance_km"],
            r["nearest_supermarket"],
            "YES -- double check this one" if flagged else "",
        ])
        if flagged:
            row_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = flag_fill

    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))] + [
                len(str(row[col_idx - 1])) for row in
                [[r["name"], r["province"], r["population"], r["distance_km"],
                  r["nearest_supermarket"],
                  "YES -- double check this one" if r.get("flagged_for_review", False) else ""]
                 for r in sorted_results]
            ]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.freeze_panes = "A2"
    wb.save(output_path)

    flagged_count = sum(1 for r in results if r.get("flagged_for_review", False))
    print(f"[INFO] Spreadsheet written: {output_path} ({len(results)} rows, "
          f"{flagged_count} flagged for manual review)")