"""
export_spreadsheet.py
---------------------
Exports the final results list to a clean, formatted .xlsx spreadsheet.

Rows at or beyond config.DISTANCE_REVIEW_THRESHOLD_KM are highlighted in
amber and marked in a dedicated column: these are results worth a manual
look rather than being silently trusted.

Rows are built once and reused for both writing and column-width measurement,
rather than being reconstructed per column.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Town",
    "Province",
    "Population",
    "Residents 65+",
    "% Aged 65+",
    "Aging Index",
    "Distance to Nearest Supermarket (km)",
    "Km Beyond Threshold",
    "Vulnerability Score",
    "Nearest Supermarket",
    "Flagged for Review",
]

HEADER_FILL = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FLAG_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

MAX_COLUMN_WIDTH = 55  # stop one long shop name from creating a 200-char column


def _build_row(result):
    """Turns one result dict into the ordered list of cell values for a row."""
    flagged = result.get("flagged_for_review", False)
    return [
        result["name"],
        result["province"],
        result["population"],
        result.get("population_65plus"),
        result.get("pct_65plus"),
        result.get("aging_index"),
        result["distance_km"],
        result.get("excess_km"),
        result.get("vulnerability"),
        result["nearest_supermarket"],
        "YES -- double check this one" if flagged else "",
    ]


def export_to_spreadsheet(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Underserved Towns"

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Sorted by vulnerability, not raw distance.
    #
    # Distance alone puts three-resident hamlets at the top -- the most extreme
    # cases, but not the largest. Vulnerability (residents 65+ x km beyond the
    # threshold) surfaces where the problem affects the most people, which is
    # what anyone acting on this list actually needs. Distance remains a column,
    # and the sheet has an auto-filter, so it can still be re-sorted freely.
    sorted_results = sorted(
        results,
        key=lambda r: (r.get("vulnerability") or 0, r["distance_km"]),
        reverse=True,
    )

    # Built once, then used for BOTH writing and width measurement.
    rows = [_build_row(r) for r in sorted_results]

    for result, row in zip(sorted_results, rows):
        ws.append(row)
        if result.get("flagged_for_review", False):
            row_idx = ws.max_row
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = FLAG_FILL

    # Single pass over the pre-built rows to size each column.
    for col_idx, header in enumerate(HEADERS, start=1):
        widest = max(
            [len(str(header))] + [len(str(row[col_idx - 1])) for row in rows]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            widest + 4, MAX_COLUMN_WIDTH
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions  # sortable/filterable straight out of the box
    wb.save(output_path)

    flagged_count = sum(1 for r in results if r.get("flagged_for_review", False))
    print(f"[INFO] Spreadsheet written: {output_path} ({len(results)} rows, "
          f"{flagged_count} flagged for manual review)")
